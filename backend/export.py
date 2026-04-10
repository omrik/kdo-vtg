import csv
import io
from datetime import datetime
from typing import Optional
from sqlalchemy.orm import Session
from sqlalchemy import desc

from backend.database import Video, ScanJob


def videos_to_csv(videos: list[Video]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        "Filename",
        "Path",
        "Resolution",
        "Width",
        "Height",
        "Duration (s)",
        "FPS",
        "Codec",
        "Bitrate",
        "Camera Type",
        "Date Created",
        "File Size (MB)",
        "Tags",
        "YOLO Enabled",
        "Scanned At",
    ])
    
    for video in videos:
        writer.writerow([
            video.filename,
            video.filepath,
            video.resolution or "",
            video.width or "",
            video.height or "",
            f"{video.duration:.2f}" if video.duration else "",
            f"{video.fps:.2f}" if video.fps else "",
            video.codec or "",
            video.bitrate or "",
            video.camera_type or "",
            video.date_created.strftime("%Y-%m-%d %H:%M:%S") if video.date_created else "",
            f"{video.file_size / (1024*1024):.2f}" if video.file_size else "",
            ", ".join(video.tags) if video.tags else "",
            "Yes" if video.yolo_enabled else "No",
            video.created_at.strftime("%Y-%m-%d %H:%M:%S") if video.created_at else "",
        ])
    
    return output.getvalue()


def videos_to_excel(videos: list[Video]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Video Metadata"
        
        headers = [
            "Filename",
            "Path",
            "Resolution",
            "Width",
            "Height",
            "Duration (s)",
            "FPS",
            "Codec",
            "Bitrate",
            "Camera Type",
            "Date Created",
            "File Size (MB)",
            "Tags",
            "YOLO Enabled",
            "Scanned At",
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, video in enumerate(videos, 2):
            ws.cell(row=row_idx, column=1, value=video.filename)
            ws.cell(row=row_idx, column=2, value=video.filepath)
            ws.cell(row=row_idx, column=3, value=video.resolution or "")
            ws.cell(row=row_idx, column=4, value=video.width or "")
            ws.cell(row=row_idx, column=5, value=video.height or "")
            ws.cell(row=row_idx, column=6, value=round(video.duration, 2) if video.duration else "")
            ws.cell(row=row_idx, column=7, value=round(video.fps, 2) if video.fps else "")
            ws.cell(row=row_idx, column=8, value=video.codec or "")
            ws.cell(row=row_idx, column=9, value=video.bitrate or "")
            ws.cell(row=row_idx, column=10, value=video.camera_type or "")
            ws.cell(row=row_idx, column=11, value=video.date_created.strftime("%Y-%m-%d %H:%M:%S") if video.date_created else "")
            ws.cell(row=row_idx, column=12, value=round(video.file_size / (1024*1024), 2) if video.file_size else "")
            ws.cell(row=row_idx, column=13, value=", ".join(video.tags) if video.tags else "")
            ws.cell(row=row_idx, column=14, value="Yes" if video.yolo_enabled else "No")
            ws.cell(row=row_idx, column=15, value=video.created_at.strftime("%Y-%m-%d %H:%M:%S") if video.created_at else "")
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    except ImportError:
        return videos_to_csv(videos).encode('utf-8')


def get_latest_scan(db: Session, folder_path: Optional[str] = None) -> Optional[ScanJob]:
    query = db.query(ScanJob)
    if folder_path:
        query = query.filter(ScanJob.folder_path == folder_path)
    return query.order_by(desc(ScanJob.created_at)).first()
